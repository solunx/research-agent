"""
Generic filesystem observer — list / inspect workspace paths.

No domain parsers (no "xlsx because task 07"). Code only:
  - list candidate roots
  - report path existence, suffix, size
  - optional light open: sheet names via openpyxl if installed, else first bytes / csv header

Used by acquisition OPEN_FILE and by tasks that name local sources.
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


DEFAULT_ROOTS = ("inputs", "runs", ".")


def list_paths(
    roots: list[str] | tuple[str, ...] | None = None,
    *,
    patterns: list[str] | None = None,
    max_entries: int = 80,
) -> dict[str, Any]:
    """
    List files under roots (relative to cwd). patterns e.g. ['*.xlsx', '*.csv'].
    """
    roots = roots or DEFAULT_ROOTS
    patterns = patterns or ["*"]
    found: list[dict[str, Any]] = []
    tried: list[str] = []

    for root in roots:
        p = Path(root)
        tried.append(str(p))
        if not p.exists():
            continue
        if p.is_file():
            found.append(_file_meta(p))
            continue
        for pat in patterns:
            for child in sorted(p.rglob(pat)):
                if child.is_file():
                    found.append(_file_meta(child))
                if len(found) >= max_entries:
                    return {
                        "ok": True,
                        "roots_tried": tried,
                        "count": len(found),
                        "files": found,
                        "truncated": True,
                    }
    return {
        "ok": True,
        "roots_tried": tried,
        "count": len(found),
        "files": found,
        "truncated": False,
    }


def _file_meta(path: Path) -> dict[str, Any]:
    try:
        st = path.stat()
        size = st.st_size
    except OSError:
        size = None
    return {
        "path": str(path),
        "name": path.name,
        "suffix": path.suffix.lower(),
        "size_bytes": size,
    }


def inspect_path(path: str | Path) -> dict[str, Any]:
    """
    Inspect one path: existence, meta, and light content summary.
    Spreadsheet: sheet names if openpyxl available; else note unavailable.
    CSV/TSV: header row + row estimate.
    Other text: first ~500 chars.
    """
    p = Path(path)
    if not p.exists():
        return {
            "ok": False,
            "exists": False,
            "path": str(p),
            "error": "path does not exist",
        }
    if not p.is_file():
        return {
            "ok": False,
            "exists": True,
            "path": str(p),
            "error": "not a file",
            "is_dir": p.is_dir(),
        }

    meta = _file_meta(p)
    meta["ok"] = True
    meta["exists"] = True
    suffix = meta["suffix"]

    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        meta.update(_inspect_xlsx(p))
    elif suffix in (".csv", ".tsv"):
        meta.update(_inspect_csv(p, delim="," if suffix == ".csv" else "\t"))
    elif suffix in (".txt", ".md", ".json", ".jsonl", ".log"):
        meta.update(_inspect_text(p))
    else:
        meta["content_summary"] = {
            "note": f"no light inspector for suffix {suffix}; path confirmed",
        }
    return meta


def _inspect_xlsx(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return {
            "content_summary": {
                "format": "xlsx",
                "sheets": None,
                "note": "openpyxl not installed; file exists but sheets not read",
            }
        }
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        first_headers: list[str] = []
        row_estimate = None
        if names:
            ws = wb[names[0]]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
                first_headers = [str(c) if c is not None else "" for c in header]
            except StopIteration:
                header = None
            n = 0
            for _ in rows:
                n += 1
                if n >= 100000:
                    break
            row_estimate = n
        wb.close()
        priceish = [
            h
            for h in first_headers
            if any(
                tok in h.lower()
                for tok in ("price", "prijs", "amount", "cost", "total", "fee", "bedrag")
            )
        ]
        return {
            "content_summary": {
                "format": "xlsx",
                "sheets": names,
                "first_sheet_headers": first_headers,
                "first_sheet_data_row_estimate": row_estimate,
                "price_or_amount_columns": priceish,
            }
        }
    except Exception as e:
        return {
            "content_summary": {
                "format": "xlsx",
                "error": f"{type(e).__name__}: {e}",
            }
        }


def _inspect_csv(path: Path, delim: str = ",") -> dict[str, Any]:
    try:
        with path.open(newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delim)
            try:
                header = next(reader)
            except StopIteration:
                return {
                    "content_summary": {
                        "format": "csv",
                        "headers": [],
                        "data_row_estimate": 0,
                    }
                }
            n = 0
            for _ in reader:
                n += 1
                if n >= 100000:
                    break
        priceish = [
            h
            for h in header
            if any(
                tok in h.lower()
                for tok in ("price", "prijs", "amount", "cost", "total", "fee", "bedrag")
            )
        ]
        return {
            "content_summary": {
                "format": "csv",
                "headers": header,
                "data_row_estimate": n,
                "price_or_amount_columns": priceish,
            }
        }
    except Exception as e:
        return {
            "content_summary": {
                "format": "csv",
                "error": f"{type(e).__name__}: {e}",
            }
        }


def _inspect_text(path: Path, max_chars: int = 500) -> dict[str, Any]:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        return {
            "content_summary": {
                "format": "text",
                "chars": len(text),
                "preview": text[:max_chars],
            }
        }
    except Exception as e:
        return {
            "content_summary": {
                "format": "text",
                "error": f"{type(e).__name__}: {e}",
            }
        }
